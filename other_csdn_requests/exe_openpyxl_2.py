import openpyxl

# 打开一个已经存在的工作簿
data = openpyxl.load_workbook('love.xlsx')
print(type(data)) # <class 'openpyxl.workbook.workbook.Workbook'>
print(data) # <openpyxl.workbook.workbook.Workbook object at 0x000001F3D3D3A7F0>

# 获取
sheet = data.active

# 获取单元格
a1 = sheet['A1'] # 获取A1单元格
print(a1) # <Cell '努埃拉'.A1>

# 获取单元格的值    
content = a1.value
print(content)

# 调用表的cell方法时，可传入整数参数，表示行和列
content2 = sheet.cell(2, 2).value
print(content2)

# 获取单元格的行和列
row = a1.row
print('行:', row) # 行: 1

column = a1.column
print('列:', column) # 列: 1

coordinate = a1.coordinate
print('坐标:', coordinate) # 坐标: A1
print('-' * 50)

# 获取第二列的所有内容
row_num = sheet.max_row # 获取最大行数
for i in range(1, row_num + 1):
    cell = sheet.cell(row, 2)
    print(cell.value)