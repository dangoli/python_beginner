import openpyxl
from openpyxl.utils import get_column_letter

# 打开一个已经存在的工作簿
data = openpyxl.load_workbook('love.xlsx')

# 获取
sheet = data.active

# 1.
max_column = sheet.max_column # 获取最大列数
column = get_column_letter(max_column) # 获取最大列数对应的字母
# 获取第一行的所有内容
row2 = sheet['A1':'%s1'%column] 
print(row2)
for row_cells in row2:
    for cell in row_cells:
        print(cell.coordinate, cell.value)
print('-' * 50)

# 2.获取第二列的所有内容
max_row = sheet.max_row # 获取最大行数
columnB = sheet['B1':'B%d'%max_row] # 获取第二列的所有内容
for column_cells in columnB:
    for cell in column_cells:
        print(cell.coordinate, cell.value)
print('-' * 50)

# 3.获取矩形区域的所有内容
cell_tuples = sheet['A1':'C3'] # 获取矩形区域的所有内容
print(cell_tuples)
for cells in cell_tuples:
    for cell in cells:
        print(cell.coordinate, cell.value)