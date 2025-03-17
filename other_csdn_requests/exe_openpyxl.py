import openpyxl

# 打开给定的Excel文件
wb = openpyxl.load_workbook('love.xlsx')
# 获取工作表
sheet = wb.active
# 获取单元格    行号从1开始，列号从A开始
cell = sheet['A1']
# 获取单元格的值
print(cell.value)
# 获取单元格的行和列
print(cell.row, cell.column)
# 获取单元格的坐标
print(cell.coordinate)
# 获取单元格的值
print(sheet.cell(row=1, column=2).value)
# 获取工作表的最大行和最大列
print(sheet.max_row, sheet.max_column)
# 获取工作表的标题
print(sheet.title)
# 获取工作表的所有行
for row in sheet.iter_rows():
    for cell in row:
        print(cell.value, end=' ')
    print()
# 获取工作表的所有列
for column in sheet.iter_cols():
    for cell in column:
        print(cell.value, end=' ')
    print()
print('-' * 50)
# 获取所有表的名称
sheetnames = wb.sheetnames
print(sheetnames)
print('-' * 50)
print(sheet) # <Worksheet "努埃拉">

# 根据名称获取工作表
sheet2 = wb['啵丝猫'] # <Worksheet "啵丝猫">
print(sheet2)

# 获取工作表的标题
sheet_name1 = sheet.title # <class 'str'>
sheet_name2 = sheet2.title
print(sheet_name1, sheet_name2, type(sheet_name1), type(sheet_name2))