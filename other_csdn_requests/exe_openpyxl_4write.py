import openpyxl

# 创建一个工作簿
w_data = openpyxl.Workbook()

# 获取工作簿的所有工作表
print(w_data.sheetnames) # ['Sheet']
# 默认的工作表名是'Sheet'，可以通过title属性修改

# 获取工作簿的活动工作表
sheet_active = w_data.active

# 修改工作表的标题
sheet_active.title = 'first_table'

# 保存工作簿
# w_data.save(filename='data/info.xlsx')

# 创建一个新的工作表
w_data.create_sheet()
w_data.create_sheet()
print(w_data.sheetnames) # ['first_table', 'Sheet', 'Sheet1']

w_data.create_sheet('second_table')

# 在指定位置创建一个新的工作表，默认是最后一个
w_data.create_sheet('third_table', 1)

print(w_data.sheetnames) # ['first_table', 'third_table', 'Sheet', 'Sheet1', 'second_table']

# 删除一个工作表
w_data.remove(w_data['Sheet1'])
# w_data.save(filename='data/info.xlsx')
w_data.save(filename='info.xlsx')

# 写入数据
w_data = openpyxl.load_workbook('info.xlsx')
sheet = w_data['first_table'] # 获取工作表

sheet['A1'] = '姓名'
sheet['B1'] = '年龄'
sheet['C1'] = '性别'

# 写入多行数据
data = [
    ('张三', 18, '男'),
    ('李四', 20, '女'),
    ('王五', 22, '男')
]

for i in range(2, 5):
    sheet[f'A{i}'] = data[i-2][0]
    sheet[f'B{i}'] = data[i-2][1]
    sheet[f'C{i}'] = data[i-2][2]

w_data.save(filename='info.xlsx')